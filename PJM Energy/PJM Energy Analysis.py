# -*- coding: utf-8 -*-
"""
Created on Fri Sep 19 08:00:13 2025

@author: EBUNOLUWASIMI
"""

# Import libraries
import pandas as pd
import matplotlib.pyplot as plt
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from io import StringIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

#Log Script
with open(r"C:\Scripts\log.txt", "a") as f:
    f.write(f"Task ran at {pd.Timestamp.now()}\n")
    
#Set output file path
info_path = r"C:\Scripts\report"
save_path = r"C:\Scripts\report\analysis.xlsx"
plot_path = r"C:\Scripts\plots"

#Set buffer string
buffer = StringIO()

#Load dataset, parse datetime
df = pd.read_csv(r"C:\Users\EBUNOLUWASIMI\Dropbox\GM\Data Analytics\Portfolios\LinkedIn Learning\Skillup\pjm_hourly_est.csv",parse_dates=['Datetime'])

#View dataset
print('Dataset: \n',df,'\n')
    
######## Preprocess Data for Analysis ################

#sort time series by datetime, extract Date, Weekday, Hour in separate columns and set date as index
df = df.sort_values('Datetime').reset_index(drop=True)
df['Date'] = df['Datetime'].dt.date
df['Year'] = df['Datetime'].dt.year
df['Month'] = df['Datetime'].dt.month_name()
df['MonthYear'] = df['Month'] + " " + df['Year'].astype(str)
df['Weekday'] = df['Datetime'].dt.day_name()
df['Hour'] = df['Datetime'].dt.hour

#Calculate PJM Total Loads across all regions
df['PJM_Load'] = df[['AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW','PJM_Load']].sum(axis=1)

#view preprocessed dataset
print('Preprocessed Dataset: \n',df,'\n')

############# Analyze preprocessed data ############################

#Explore dataset with Descriptives
""" Diagnostics """
info = df.info(buf=buffer)
info_str = buffer.getvalue()
with open(info_path + '\df_info.txt', 'w') as f:
    f.write(info_str)
    
describe = df.describe().round(1)
print("Data Information: \n",df.info(),'\n')
print('Descriptives: \n',describe,'\n')

#Plot hourly energy load
plt.figure(figsize=(30, 10))
plt.plot(df['PJM_Load'])
plt.title("Hourly Load")
plt.xlabel("Datetime")
plt.ylabel("Total Load")
plt.xticks(rotation=60)
plt.tight_layout()
plt.savefig(plot_path + '\Raw Data.png')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
with pd.ExcelWriter(save_path, engine='xlsxwriter') as writer:
    describe.to_excel(writer, sheet_name='Descriptives', index=True)
    workbook = writer.book
    worksheet = writer.sheets['Descriptives']
    worksheet.insert_image('S2', plot_path + '\Raw Data.png')

#Daily Load peak hour for each date
daily_peak_hours = (df.loc[df.groupby('Date')['PJM_Load'].idxmax(), ['Datetime','Date', 'Weekday', 'Hour','PJM_Load']].rename(columns={'Hour':'Peak Hour'}))

print("daily load peak hour for each weekday: \n")
print('Daily Load Peak Hour: \n',daily_peak_hours,'\n')

#Plot Daily load peak hour by Date
plt.scatter(daily_peak_hours['Datetime'], daily_peak_hours['Peak Hour'])
plt.title("Daily Load Peak Hour")
plt.xlabel("Datetime")
plt.ylabel("load Peak Hour")
plt.xticks(rotation=60)
plt.tight_layout()
plt.savefig(plot_path + '\Daily Load Peak Hour.png', bbox_inches='tight')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Daily Load Peak Hour'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]
    
img = Image('Daily Load Peak Hour.png')
ws.add_image(img, 'J2')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    daily_peak_hours.to_excel(writer, sheet_name = Sheet_Name, index=True)

# Sort weekdays in natural order
weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
daily_peak_hours['Weekday'] = pd.Categorical(daily_peak_hours['Weekday'],
    categories=weekday_order, ordered=True)
daily_peak_hours = daily_peak_hours.sort_values('Weekday')

#Plot daily peak hour by weekday
plt.scatter(daily_peak_hours['Weekday'], daily_peak_hours['Peak Hour'])
plt.title("Daily Peak Hour of Energy Consumption by Weekday")
plt.xlabel("Weekday")
plt.ylabel("Peak Hour")
plt.xticks(rotation=30)
plt.tight_layout()
plt.savefig(plot_path + '\Daily Peak Hour of Energy Consumption by Weekday.png', bbox_inches='tight')
plt.show()
plt.close()

#Insert plot into file
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Daily Load Peak Hour'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]
    
img = Image(plot_path + '\Daily Peak Hour of Energy Consumption by Weekday.png')
ws.add_image(img, 'J30')
wb.save(excel_file)

#Average Load peak hours across weeks for each weekday
avg_peak_hour_by_weekday = (daily_peak_hours.groupby('Weekday')['Peak Hour']
        .mean()
        .reset_index()
        .rename(columns={'Peak Hour':'Average_Peak_Hour'}))

# Sort weekdays in natural order
avg_peak_hour_by_weekday['Weekday'] = pd.Categorical(avg_peak_hour_by_weekday['Weekday'],
    categories=weekday_order, ordered=True)
avg_peak_hour_by_weekday = avg_peak_hour_by_weekday.sort_values('Weekday')

print("Average daily load peak hour for each weekday: \n")
print('Avg Peak Hour by Weekday \n',avg_peak_hour_by_weekday,'\n')

#Plot avg Peak hour by weekday
plt.bar(avg_peak_hour_by_weekday['Weekday'], avg_peak_hour_by_weekday['Average_Peak_Hour'])
plt.title("Avg Daily Load Peak Hour by Weekday")
plt.xlabel("Weekday")
plt.ylabel("Average Peak Hour")
plt.xticks(rotation=30)
plt.tight_layout()
plt.savefig(plot_path + '\Avg Daily Load Peak Hour by Weekday.png', bbox_inches='tight')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Avg Daily Load Peak Hour by Weekday'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]
    
img = Image(plot_path + '\Avg Daily Load Peak Hour by Weekday.png')
ws.add_image(img, 'F2')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    avg_peak_hour_by_weekday.to_excel(writer, sheet_name = Sheet_Name, index=False)

#Calculate Total Daily Load over time
Daily_Load = (df.groupby('Date')[['AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW','PJM_Load']]
    .sum()
    .reset_index()
    .rename(columns={'PJM_Load':'Total Daily Load'}))
print('Total Daily Load \n',Daily_Load,'\n')

#Plot Total Daily Load over time
plt.figure(figsize=(10, 5))
plt.plot(Daily_Load['Date'], Daily_Load['Total Daily Load'])
plt.title("Total Daily Load by Date")
plt.xlabel("Date")
plt.ylabel("Load")
plt.xticks(rotation=60)
plt.tight_layout()
plt.savefig(plot_path + '\Daily Load.png', bbox_inches='tight')
plt.show()
plt.close()

#Plot Daily Load over time by ISO
plt.figure(figsize=(15, 5))
plt.plot(Daily_Load['Date'], Daily_Load[['AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW']],label= ['AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW'])
plt.title("Daily Load by ISO")
plt.xlabel("Date")
plt.ylabel("Load")
plt.xticks(rotation=60)
plt.tight_layout()
plt.legend(loc='best',title='ISO')
plt.savefig(plot_path + '\Daily Load by ISO.png', bbox_inches='tight')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Daily Load'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]
    
img = Image(plot_path + '\Daily Load.png')
ws.add_image(img, 'P2')
img = Image(plot_path + '\Daily Load by ISO.png')
ws.add_image(img, 'P30')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    Daily_Load.to_excel(writer, sheet_name = Sheet_Name, index=False)

#Calculate Monthly Energy Demand over time
Monthly_Load = (df.groupby(['Year','Month','MonthYear'])[['PJM_Load','AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW']]
    .sum()
    .reset_index()
    .rename(columns={'PJM_Load':'Monthly Load'}))

# Sort in natural order
month_order = ["January","February", "March", "April", "May", "June","July","August","September","October","November","December"]
Monthly_Load['Month'] = pd.Categorical(Monthly_Load['Month'],
    categories=month_order, ordered=True)
Monthly_Load = Monthly_Load.sort_values(['Year','Month'])
print('Total monthly Load \n',Monthly_Load,'\n')

#Plot MOnthly Load
plt.figure(figsize=(30, 10))
plt.plot(Monthly_Load['MonthYear'], Monthly_Load['Monthly Load'])
plt.title("Total Monthly Energy Demand")
plt.xlabel("Month Name")
plt.ylabel("Total Energy Demand")
#plt.xticks([])
plt.xticks(rotation=60)
plt.tight_layout()
plt.legend(loc='best',title='Legend')
plt.savefig(plot_path + '\Monthly Load.png', bbox_inches='tight')
plt.show()
plt.close()

plt.figure(figsize=(30, 10))
plt.plot(Monthly_Load['MonthYear'], Monthly_Load[['AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW']])
plt.title("Monthly Energy Demand by ISO")
plt.xlabel("Month Name")
plt.ylabel("Total Energy Demand")
#plt.xticks([])
plt.xticks(rotation=60)
plt.tight_layout()
plt.legend(loc='best',title='Legend')
plt.savefig(plot_path + '\Monthly Load by ISO.png', bbox_inches='tight')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Monthly Load'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]
    
img = Image(plot_path + '\Monthly Load.png')
ws.add_image(img, 'P2')
img = Image(plot_path + '\Monthly Load by ISO.png')
ws.add_image(img, 'P30')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    Monthly_Load.to_excel(writer, sheet_name = Sheet_Name, index=False)

#Determine the seasons of the year
datevalue = df['Date']
def get_season(datevalue):
    m = datevalue.month
    if m in [12, 1, 2]:
        return "Winter"
    elif m in [3, 4, 5]:
        return "Spring"
    elif m in [6, 7, 8]:
        return "Summer"
    else:
        return "Autumn"

df["Season"] = df["Date"].apply(get_season)
df["Season Year"] = df["Season"]+ " " + df['Year'].astype(str)

#Calculate Seasonal Energy Demand over time
Seasonal_Load = (df.groupby(['Year','Season','Season Year'])[['PJM_Load','AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW']]
    .sum()
    .reset_index()
    .rename(columns={'PJM_Load':'Seasonal Demand'}))

#Sort in natural order of seasons
season_order = ["Winter","Spring", "Summer", "Autumn"]
Seasonal_Load['Season'] = pd.Categorical(Seasonal_Load['Season'],
    categories=season_order, ordered=True)
Seasonal_Load = Seasonal_Load.sort_values(['Year','Season'])
print(Seasonal_Load)

#Plot Total Seasonal Load
plt.figure(figsize=(30, 10))
plt.plot(Seasonal_Load['Season Year'], Seasonal_Load['Seasonal Demand'])
plt.title("Total Seasonal Load")
plt.xlabel("Season Year")
plt.ylabel("Load")
#plt.xticks([])
plt.xticks(rotation=60)
plt.tight_layout()
plt.savefig(plot_path + '\Seasonal Load.png',bbox_inches='tight')
plt.show()
plt.close()

#Plot Seasonal Demands by ISO
plt.figure(figsize=(30, 10))
plt.plot(Seasonal_Load['Season Year'], Seasonal_Load[['AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW']],label= ['AEP','COMED','DAYTON','DEOK','DOM','DUQ','EKPC','FE','NI','PJME','PJMW'])
plt.title("Seasonal Energy Demand by ISO")
plt.xlabel("Season Year")
plt.ylabel("Energy Demand")
#plt.xticks([])
plt.xticks(rotation=60)
plt.tight_layout()
plt.legend(loc='best',title='ISO')
plt.savefig(plot_path + '\Seaonal Load by ISO.png', bbox_inches='tight')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Seasonal Load'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]
    
img = Image(plot_path + '\Seasonal Load.png')
ws.add_image(img, 'P2')
img = Image(plot_path+'\Seaonal Load by ISO.png')
ws.add_image(img, 'P30')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    Seasonal_Load.to_excel(writer, sheet_name = Sheet_Name, index=False)

######Conduct Demand Forecast for the next one year########

#Split series into training and test set
Timeseries = pd.DataFrame(Daily_Load[['Date','Total Daily Load']])
Timeseries['Date'] = pd.to_datetime(Timeseries['Date'])
train = Timeseries[Timeseries['Date'] <= pd.Timestamp('2017-08-03')]
test = Timeseries[Timeseries['Date'] > pd.Timestamp('2017-08-03')]
print('Train Set: \n',train,'\n','Test Set: \n',test,'\n')

#Configure dataset as monthly time series
train = train.set_index('Date')
train = train.asfreq('ME')
train['Total Daily Load'] = train['Total Daily Load'].interpolate()

test = test.set_index('Date')
test = test.asfreq('ME')
test['Total Daily Load'] = test['Total Daily Load'].interpolate()
print('Interpolated Train Series \n',train, '\n Interpolated Test Series \n',test,'\n')

#Explore training series
plt.figure(figsize=(30, 10))
train['Total Daily Load'].plot(title="Monthly Time Series")
plt.show()

#Test for stationarity 
testres = adfuller(train)
adf_statistic = testres[0]
p_value = testres[1]
critical_values = testres[4]

print(f"ADF Statistic: {adf_statistic}")
print(f"p-value: {p_value}")
print("Critical Values:")
for key, value in critical_values.items():
    print(f"   {key}: {value}")

if p_value >= 0.05:
    print('Series is non-stationary. \n Seasonal differentiation in progress \n')
    
    #perform seasonal differencing
    seasonally_diff = train.diff(12).dropna()
    
    # Test stationarity again
    testres2 = adfuller(seasonally_diff)
    print(f"After Differencing - p-value: {testres2[1]}")
else:
    print('Stationarity is confirmed!')

#Plot series
if p_value >= 0.05:
    plt.figure(figsize=(30,10))
    plt.plot(train.index,train['Total Daily Load'],label="Original Series")
    plt.plot(seasonally_diff.index,seasonally_diff['Total Daily Load'],label="First Differenced Series",color='red')
    plt.title('Monthly Load')
    plt.tight_layout()
    plt.legend()
    plt.savefig(plot_path + '\Monthly Load.png',bbox_inches='tight')
    plt.show()
    plt.close()
else:
    plt.figure(figsize=(30,10))
    plt.plot(train.index,train['Total Daily Load'],label="Original Series")
    plt.title('Monthly Load')
    plt.tight_layout()
    plt.legend()
    plt.savefig(plot_path + '\Monthly Load.png',bbox_inches='tight')
    plt.show()
    plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Stationarity'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]

img = Image(plot_path + '\Monthly Load.png')
ws.add_image(img, 'P2')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    pd.DataFrame(testres).to_excel(writer, sheet_name = Sheet_Name, index=False)
    pd.DataFrame(testres2).to_excel(writer, sheet_name = Sheet_Name, index=False)

#Fit ARIMA model
arima = ARIMA(train['Total Daily Load'], order=(1, 1, 1))
arima_fit = arima.fit()
arima_summary = arima_fit.summary()

print(arima_summary)

#Explore residuals
arima_fit.plot_diagnostics(figsize=(10,6))
plt.savefig(plot_path + '\Arima Diagnostics.png',bbox_inches='tight')
plt.show()
plt.close()

# Forecast the next 12 periods
arima_forecast = arima_fit.forecast(steps=12)
arima_forecast.index = pd.date_range(train.index[-1] + pd.offsets.MonthEnd(),periods=12, freq='ME')
print('12 Months Forecast: \n',arima_forecast,'\n')

# Plot forecast
plt.figure(figsize=(30,10))
plt.plot(train.index, train['Total Daily Load'], label='Observed')
plt.plot(arima_forecast.index, arima_forecast, label='Forecast', color='red')
plt.legend()
plt.savefig(plot_path + '\ARIMA Forecast.png',bbox_inches='tight')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='ARIMA'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]

img = Image(plot_path + '\Arima Diagnostics.png')
ws.add_image(img, 'P2')
img = Image(plot_path + '\Arima Forecast.png')
ws.add_image(img, 'P50')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    arima_forecast.to_excel(writer, sheet_name = Sheet_Name, index=False)


#Fit Holt-Winters Exponential Smoothening model
exp_smth = train
exp_model = ExponentialSmoothing(
    exp_smth['Total Daily Load'],
    trend='add',          
    seasonal='add',       
    seasonal_periods=12).fit()

# Fitted values
exp_smth['HW_Fitted'] = exp_model.fittedvalues

# Forecast next 12 periods
exp_forecast = exp_model.forecast(12)

# --- Plot ---
plt.figure(figsize=(30,10))
plt.plot(exp_smth.index, exp_smth['Total Daily Load'], label='Original', marker='o')
plt.plot(exp_smth.index, exp_smth['HW_Fitted'], label='HW Fitted', linestyle='--')
plt.plot(pd.date_range(exp_smth.index[-1] + pd.offsets.MonthEnd(), periods=12, freq='M'),
         exp_forecast, label='Forecast', marker='x', color='red')
plt.title('Holt-Winters Exponential Smoothing')
plt.xlabel('Date')
plt.ylabel('Monthly Load')
plt.legend()
plt.tight_layout()
plt.savefig(plot_path + '\Holt-Winters Exponential Smoothing.png',bbox_inches='tight')
plt.show()
plt.close()

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Holt-Winters Exp Smoothing'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]

img = Image(plot_path + '\Holt-Winters Exponential Smoothing.png')
ws.add_image(img, 'P2')
wb.save(excel_file)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    exp_smth.to_excel(writer, sheet_name = Sheet_Name, index=False)
    exp_forecast.to_excel(writer, sheet_name = Sheet_Name, index=False)

#Calculating Forecast Residuals
arima_residuals = test['Total Daily Load'] - arima_forecast
arima_msr = (arima_residuals**2).mean()
arima_sterror = arima_msr**0.5

exp_residuals = test['Total Daily Load'] - exp_forecast
exp_msr = (exp_residuals**2).mean()
exp_sterror = exp_msr**0.5
print('Residuals: \n ARIMA \n',arima_residuals,'\n  Holt-Winters Exponential Smoothing: \n',exp_residuals,'\n')

model_diag = {'Model':['ARIMA','H_W Exp_Smth'],'MSE':[arima_msr,exp_msr],'Std. Error':[arima_sterror,exp_sterror]}
model_diag = pd.DataFrame(model_diag)

print('Models Diagnostics: \n',model_diag,'\n')

#Write DataFrame to Excel and insert plot
excel_file = save_path
wb = load_workbook(save_path)
Sheet_Name='Residuals'
if Sheet_Name not in wb.sheetnames:
    ws = wb.create_sheet(Sheet_Name)
else:
    ws = wb[Sheet_Name]

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    test.to_excel(writer, sheet_name = Sheet_Name, index=False)
    arima_residuals.to_excel(writer, sheet_name = Sheet_Name, index=False)
    exp_residuals.to_excel(writer, sheet_name = Sheet_Name, index=False)
    model_diag.to_excel(writer, sheet_name = Sheet_Name, index=False)








